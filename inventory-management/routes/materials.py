"""
Materials master data management routes
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Material, InventoryLevel, Category, Provider, Batch, Location, Bin
from sqlalchemy import func, or_
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime

bp = Blueprint('materials', __name__)


@bp.route('/')
@login_required
def index():
    """List all materials"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    category = request.args.get('category', '', type=str)
    provider_id = request.args.get('provider_id', '', type=str)
    unit_of_measure = request.args.get('unit_of_measure', '', type=str)
    active_filter = request.args.get('active', '', type=str)

    query = Material.query

    # Search filter
    if search:
        query = query.filter(
            or_(
                Material.name.ilike(f'%{search}%'),
                Material.description.ilike(f'%{search}%')
            )
        )

    # Category filter
    if category:
        query = query.filter(Material.category == category)

    # Provider filter
    if provider_id:
        query = query.filter(Material.provider_id == int(provider_id))

    # Unit of Measure filter
    if unit_of_measure:
        query = query.filter(Material.unit_of_measure.ilike(f'%{unit_of_measure}%'))

    # Active/Inactive filter
    if active_filter == 'active':
        query = query.filter(Material.active == True)
    elif active_filter == 'inactive':
        query = query.filter(Material.active == False)

    # Get all categories for filter dropdown
    categories = db.session.query(Material.category).filter(
        Material.category.isnot(None),
        Material.category != ''
    ).distinct().order_by(Material.category).all()
    categories = [c[0] for c in categories]

    # Get all active providers for filter dropdown
    from models import Provider
    providers = Provider.query.filter_by(active=True).order_by(Provider.name).all()

    # Pagination
    pagination = query.order_by(Material.name).paginate(
        page=page, per_page=50, error_out=False
    )
    materials = pagination.items

    # Get current stock for each material
    materials_with_stock = []
    for material in materials:
        stock_qty = db.session.query(func.sum(InventoryLevel.quantity)).filter(
            InventoryLevel.material_id == material.id
        ).scalar() or 0
        materials_with_stock.append((material, stock_qty))

    return render_template('materials/index.html',
                          materials=materials_with_stock,
                          pagination=pagination,
                          search=search,
                          category=category,
                          categories=categories,
                          provider_id=provider_id,
                          providers=providers,
                          unit_of_measure=unit_of_measure,
                          active_filter=active_filter)


@bp.route('/warehouse')
@login_required
def warehouse_view():
    """Warehouse view with material and item dimensions for warehouse managers"""
    from models import Item, Client

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    item_type = request.args.get('type', '', type=str)  # 'material' or 'item'
    category_id = request.args.get('category_id', '', type=str)
    provider_id = request.args.get('provider_id', '', type=str)
    client_id = request.args.get('client_id', '', type=str)
    location_id = request.args.get('location_id', '', type=str)
    has_dimensions = request.args.get('has_dimensions', '', type=str)
    active_filter = request.args.get('active', 'active', type=str)

    # Collect all items (materials and items) with their inventory
    all_items = []

    # Query materials if not filtered to items only
    if item_type != 'item':
        material_query = Material.query

        if search:
            material_query = material_query.filter(
                or_(
                    Material.name.ilike(f'%{search}%'),
                    Material.description.ilike(f'%{search}%')
                )
            )

        if category_id:
            material_query = material_query.filter(Material.category_id == int(category_id))

        if provider_id:
            material_query = material_query.filter(Material.provider_id == int(provider_id))

        if has_dimensions == 'yes':
            material_query = material_query.filter(
                or_(
                    Material.diameter.isnot(None),
                    Material.length.isnot(None),
                    Material.width.isnot(None),
                    Material.height.isnot(None)
                )
            )
        elif has_dimensions == 'no':
            material_query = material_query.filter(
                Material.diameter.is_(None),
                Material.length.is_(None),
                Material.width.is_(None),
                Material.height.is_(None)
            )

        if active_filter == 'active':
            material_query = material_query.filter(Material.active == True)
        elif active_filter == 'inactive':
            material_query = material_query.filter(Material.active == False)

        materials = material_query.order_by(Material.name).all()

        for material in materials:
            total_stock = db.session.query(func.sum(InventoryLevel.quantity)).filter(
                InventoryLevel.material_id == material.id
            ).scalar() or 0

            # Always get detailed location/bin information
            stock_by_location = db.session.query(
                Location.code,
                Location.name,
                Bin.bin_code,
                InventoryLevel.quantity
            ).join(
                InventoryLevel, InventoryLevel.location_id == Location.id
            ).outerjoin(
                Bin, InventoryLevel.bin_id == Bin.id
            ).filter(
                InventoryLevel.material_id == material.id
            )

            if location_id:
                stock_by_location = stock_by_location.filter(Location.id == int(location_id))

            stock_by_location = stock_by_location.order_by(Location.code, Bin.bin_code).all()

            all_items.append({
                'type': 'material',
                'item': material,
                'total_stock': total_stock,
                'stock_by_location': stock_by_location,
                'organization': material.provider.name if material.provider else None
            })

    # Query items if not filtered to materials only
    if item_type != 'material':
        item_query = Item.query

        if search:
            item_query = item_query.filter(
                or_(
                    Item.name.ilike(f'%{search}%'),
                    Item.description.ilike(f'%{search}%')
                )
            )

        if category_id:
            item_query = item_query.filter(Item.category_id == int(category_id))

        if client_id:
            item_query = item_query.filter(Item.client_id == int(client_id))

        if has_dimensions == 'yes':
            item_query = item_query.filter(
                or_(
                    Item.diameter.isnot(None),
                    Item.length.isnot(None),
                    Item.width.isnot(None),
                    Item.height.isnot(None)
                )
            )
        elif has_dimensions == 'no':
            item_query = item_query.filter(
                Item.diameter.is_(None),
                Item.length.is_(None),
                Item.width.is_(None),
                Item.height.is_(None)
            )

        if active_filter == 'active':
            item_query = item_query.filter(Item.active == True)
        elif active_filter == 'inactive':
            item_query = item_query.filter(Item.active == False)

        items = item_query.order_by(Item.name).all()

        for item in items:
            total_stock = db.session.query(func.sum(InventoryLevel.quantity)).filter(
                InventoryLevel.item_id == item.id
            ).scalar() or 0

            # Always get detailed location/bin information
            stock_by_location = db.session.query(
                Location.code,
                Location.name,
                Bin.bin_code,
                InventoryLevel.quantity
            ).join(
                InventoryLevel, InventoryLevel.location_id == Location.id
            ).outerjoin(
                Bin, InventoryLevel.bin_id == Bin.id
            ).filter(
                InventoryLevel.item_id == item.id
            )

            if location_id:
                stock_by_location = stock_by_location.filter(Location.id == int(location_id))

            stock_by_location = stock_by_location.order_by(Location.code, Bin.bin_code).all()

            all_items.append({
                'type': 'item',
                'item': item,
                'total_stock': total_stock,
                'stock_by_location': stock_by_location,
                'organization': item.client.name if item.client else None
            })

    # Sort all items by name
    all_items.sort(key=lambda x: x['item'].name)

    # Manual pagination
    total_items = len(all_items)
    per_page = 50
    start = (page - 1) * per_page
    end = start + per_page
    paginated_items = all_items[start:end]

    # Create pagination object manually
    class SimplePagination:
        def __init__(self, page, per_page, total):
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page if total > 0 else 1
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

        def iter_pages(self, left_edge=1, right_edge=1, left_current=2, right_current=2):
            last = 0
            for num in range(1, self.pages + 1):
                if (num <= left_edge or
                    (self.page - left_current - 1 < num < self.page + right_current) or
                    num > self.pages - right_edge):
                    if last + 1 != num:
                        yield None
                    yield num
                    last = num

    pagination = SimplePagination(page, per_page, total_items)

    # Get all categories for both types
    material_categories = Category.query.filter_by(category_type='material', active=True).order_by(Category.name).all()
    item_categories = Category.query.filter_by(category_type='item', active=True).order_by(Category.name).all()

    # Get providers and clients
    providers = Provider.query.filter_by(active=True).order_by(Provider.name).all()
    clients = Client.query.filter_by(active=True).order_by(Client.name).all()

    # Get locations
    locations = Location.query.filter_by(active=True).order_by(Location.code).all()

    return render_template('materials/warehouse.html',
                          items=paginated_items,
                          pagination=pagination,
                          search=search,
                          item_type=item_type,
                          category_id=category_id,
                          material_categories=material_categories,
                          item_categories=item_categories,
                          provider_id=provider_id,
                          providers=providers,
                          client_id=client_id,
                          clients=clients,
                          location_id=location_id,
                          locations=locations,
                          has_dimensions=has_dimensions,
                          active_filter=active_filter)


@bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    """Create new material"""
    if request.method == 'POST':
        try:
            # Get category_id and provider_id from form
            category_id = request.form.get('category_id')
            if category_id:
                category_id = int(category_id) if category_id else None

            provider_id = request.form.get('provider_id')
            if provider_id:
                provider_id = int(provider_id) if provider_id else None

            material = Material(
                name=request.form['name'].strip(),
                description=request.form.get('description', '').strip(),
                category=request.form.get('category', '').strip(),  # Keep legacy field for backwards compatibility
                category_id=category_id,
                provider_id=provider_id,
                unit_of_measure=request.form['unit_of_measure'].strip(),
                reorder_level=float(request.form.get('reorder_level', 0)),
                reorder_quantity=float(request.form.get('reorder_quantity', 0)),
                diameter=float(request.form.get('diameter')) if request.form.get('diameter') else None,
                length=float(request.form.get('length')) if request.form.get('length') else None,
                width=float(request.form.get('width')) if request.form.get('width') else None,
                height=float(request.form.get('height')) if request.form.get('height') else None,
                active=request.form.get('active') == 'on'
            )

            db.session.add(material)
            db.session.commit()

            flash(f'Material "{material.name}" created successfully!', 'success')
            return redirect(url_for('materials.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating material: {str(e)}', 'danger')

    # Get active material categories and providers
    categories = Category.query.filter_by(category_type='material', active=True).order_by(Category.name).all()
    providers = Provider.query.filter_by(active=True).order_by(Provider.name).all()

    return render_template('materials/new.html', categories=categories, providers=providers)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    """Edit material"""
    material = Material.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # Get category_id and provider_id from form
            category_id = request.form.get('category_id')
            if category_id:
                category_id = int(category_id) if category_id else None

            provider_id = request.form.get('provider_id')
            if provider_id:
                provider_id = int(provider_id) if provider_id else None

            material.name = request.form['name'].strip()
            material.description = request.form.get('description', '').strip()
            material.category = request.form.get('category', '').strip()  # Keep legacy field
            material.category_id = category_id
            material.provider_id = provider_id
            material.unit_of_measure = request.form['unit_of_measure'].strip()
            material.reorder_level = float(request.form.get('reorder_level', 0))
            material.reorder_quantity = float(request.form.get('reorder_quantity', 0))
            material.diameter = float(request.form.get('diameter')) if request.form.get('diameter') else None
            material.length = float(request.form.get('length')) if request.form.get('length') else None
            material.width = float(request.form.get('width')) if request.form.get('width') else None
            material.height = float(request.form.get('height')) if request.form.get('height') else None
            material.active = request.form.get('active') == 'on'

            db.session.commit()

            flash(f'Material "{material.name}" updated successfully!', 'success')
            return redirect(url_for('materials.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating material: {str(e)}', 'danger')

    # Get active material categories and providers
    categories = Category.query.filter_by(category_type='material', active=True).order_by(Category.name).all()
    providers = Provider.query.filter_by(active=True).order_by(Provider.name).all()

    return render_template('materials/edit.html', material=material, categories=categories, providers=providers)


@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    """Delete material"""
    material = Material.query.get_or_404(id)

    # Check if material has inventory
    has_inventory = InventoryLevel.query.filter_by(material_id=id).first() is not None

    if has_inventory:
        flash(f'Cannot delete material "{material.name}" - it has inventory records.', 'danger')
        return redirect(url_for('materials.index'))

    try:
        db.session.delete(material)
        db.session.commit()
        flash(f'Material "{material.name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting material: {str(e)}', 'danger')

    return redirect(url_for('materials.index'))


@bp.route('/export')
@login_required
def export():
    """Export materials to Excel with batch information"""
    materials = Material.query.order_by(Material.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    # Headers
    headers = ['Name', 'Description', 'Category', 'Unit of Measure',
               'Reorder Level', 'Reorder Quantity', 'Diameter', 'Length', 'Width', 'Height', 'Active',
               'Batch Number', 'Supplier Batch Number', 'Location Code', 'Bin Code',
               'Quantity', 'Cost per Unit', 'Received Date']
    ws.append(headers)

    # Data
    for material in materials:
        # Get first active batch for this material (if exists)
        batch = Batch.query.filter_by(
            material_id=material.id,
            status='active'
        ).order_by(Batch.received_date.desc()).first()

        # Base material data
        row_data = [
            material.name,
            material.description,
            material.category,
            material.unit_of_measure,
            material.reorder_level,
            material.reorder_quantity,
            material.diameter if material.diameter else '',
            material.length if material.length else '',
            material.width if material.width else '',
            material.height if material.height else '',
            'Yes' if material.active else 'No'
        ]

        # Add batch data if available
        if batch:
            location = Location.query.get(batch.location_id)
            bin_obj = Bin.query.get(batch.bin_id) if batch.bin_id else None
            row_data.extend([
                batch.batch_number,
                batch.supplier_batch_number or '',
                location.code if location else '',
                bin_obj.bin_code if bin_obj else '',
                batch.quantity_available,
                float(batch.cost_per_unit),
                batch.received_date.strftime('%Y-%m-%d') if batch.received_date else ''
            ])
        else:
            # Empty batch columns
            row_data.extend(['', '', '', '', '', '', ''])

        ws.append(row_data)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"materials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_data():
    """Import materials from Excel"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded.', 'danger')
            return redirect(url_for('materials.import_data'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('materials.import_data'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('materials.import_data'))

        try:
            wb = load_workbook(file)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            imported = 0
            updated = 0
            errors = []

            for idx, row in enumerate(rows, start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    name = str(row[0]).strip()
                    description = str(row[1] or '').strip()
                    category = str(row[2] or '').strip()
                    unit_of_measure = str(row[3] or 'PCS').strip()

                    # Safe float conversion
                    try:
                        reorder_level = float(row[4]) if row[4] not in [None, ''] else 0
                    except (ValueError, TypeError):
                        reorder_level = 0

                    try:
                        reorder_quantity = float(row[5]) if row[5] not in [None, ''] else 0
                    except (ValueError, TypeError):
                        reorder_quantity = 0

                    active = str(row[6] or 'Yes').lower() in ['yes', 'true', '1', 'active']

                    # Batch information (optional)
                    batch_number = str(row[7]).strip() if len(row) > 7 and row[7] not in [None, ''] else None
                    supplier_batch_number = str(row[8]).strip() if len(row) > 8 and row[8] not in [None, ''] else None
                    location_code = str(row[9]).strip() if len(row) > 9 and row[9] not in [None, ''] else None
                    bin_code = str(row[10]).strip() if len(row) > 10 and row[10] not in [None, ''] else None

                    # Safe float conversion for batch quantity and cost
                    quantity = None
                    if len(row) > 11 and row[11] not in [None, '']:
                        try:
                            quantity = float(row[11])
                        except (ValueError, TypeError):
                            pass

                    cost_per_unit = None
                    if len(row) > 12 and row[12] not in [None, '']:
                        try:
                            cost_per_unit = float(row[12])
                        except (ValueError, TypeError):
                            pass

                    received_date_str = str(row[13]).strip() if len(row) > 13 and row[13] not in [None, ''] else None

                    # Check if material exists
                    material = Material.query.filter_by(name=name).first()

                    if material:
                        # Update existing
                        material.description = description
                        material.category = category
                        material.unit_of_measure = unit_of_measure
                        material.reorder_level = reorder_level
                        material.reorder_quantity = reorder_quantity
                        material.active = active
                        updated += 1
                    else:
                        # Create new
                        material = Material(
                            name=name,
                            description=description,
                            category=category,
                            unit_of_measure=unit_of_measure,
                            reorder_level=reorder_level,
                            reorder_quantity=reorder_quantity,
                            active=active
                        )
                        db.session.add(material)
                        imported += 1

                    # Flush to get material ID for batch creation
                    db.session.flush()

                    # Create batch if batch information is provided
                    if batch_number and location_code and quantity and cost_per_unit:
                        # Find location
                        location = Location.query.filter_by(code=location_code).first()
                        if not location:
                            errors.append(f"Row {idx}: Location '{location_code}' not found")
                            continue

                        # Find bin if provided
                        bin_obj = None
                        if bin_code:
                            bin_obj = Bin.query.filter_by(location_id=location.id, bin_code=bin_code).first()
                            if not bin_obj:
                                errors.append(f"Row {idx}: Bin '{bin_code}' not found in location '{location_code}'")
                                continue

                        # Parse received date
                        received_date = datetime.utcnow()
                        if received_date_str:
                            try:
                                # Try different date formats
                                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                                    try:
                                        received_date = datetime.strptime(received_date_str, fmt)
                                        break
                                    except ValueError:
                                        continue
                            except:
                                pass  # Use default date if parsing fails

                        # Check if batch already exists
                        existing_batch = Batch.query.filter_by(batch_number=batch_number).first()
                        if not existing_batch:
                            # Create new batch
                            batch = Batch(
                                batch_number=batch_number,
                                material_id=material.id,
                                location_id=location.id,
                                bin_id=bin_obj.id if bin_obj else None,
                                quantity_original=quantity,
                                quantity_available=quantity,
                                cost_per_unit=cost_per_unit,
                                supplier_batch_number=supplier_batch_number,
                                received_date=received_date,
                                status='active'
                            )
                            db.session.add(batch)

                            # Update inventory level
                            inventory = InventoryLevel.query.filter_by(
                                material_id=material.id,
                                location_id=location.id,
                                bin_id=bin_obj.id if bin_obj else None
                            ).first()

                            if inventory:
                                inventory.quantity += quantity
                            else:
                                inventory = InventoryLevel(
                                    material_id=material.id,
                                    location_id=location.id,
                                    bin_id=bin_obj.id if bin_obj else None,
                                    quantity=quantity
                                )
                                db.session.add(inventory)

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            db.session.commit()

            if errors:
                flash(f'Import completed with errors. Imported: {imported}, Updated: {updated}. Errors: {"; ".join(errors[:5])}', 'warning')
            else:
                flash(f'Import successful! Imported: {imported}, Updated: {updated}.', 'success')

            return redirect(url_for('materials.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error importing file: {str(e)}', 'danger')

    return render_template('materials/import.html')


@bp.route('/template')
@login_required
def download_template():
    """Download Excel template for materials import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials Template"

    # Headers
    headers = ['Name', 'Description', 'Category', 'Unit of Measure',
               'Reorder Level', 'Reorder Quantity', 'Active', 'Batch Number',
               'Supplier Batch Number', 'Location Code', 'Bin Code', 'Quantity',
               'Cost per Unit', 'Received Date']
    ws.append(headers)

    # Sample data
    ws.append(['Steel Plate 10mm', 'Steel plate 10mm thickness', 'Metals', 'PCS', 100, 200, 'Yes',
               'BATCH-2025-001', 'SUP-12345', 'WH-001', 'A-01', 500, 25.50, '2025-01-15'])
    ws.append(['Plastic Resin', 'High-density plastic resin', 'Plastics', 'KG', 500, 1000, 'Yes',
               'BATCH-2025-002', 'SUP-67890', 'WH-001', 'B-02', 1500, 12.75, '2025-01-16'])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='materials_template.xlsx')
