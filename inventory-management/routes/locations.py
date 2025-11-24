"""
Locations and bins management routes
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, Location, Bin, InventoryLevel
from sqlalchemy import func
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from datetime import datetime

bp = Blueprint('locations', __name__)


@bp.route('/')
@login_required
def index():
    """List all locations"""
    # Get filter parameters
    search = request.args.get('search', '', type=str)
    location_type = request.args.get('location_type', '', type=str)
    active_filter = request.args.get('active', '', type=str)
    page = request.args.get('page', 1, type=int)

    query = Location.query

    # Search filter (by code or name)
    if search:
        from sqlalchemy import or_
        query = query.filter(
            or_(
                Location.code.ilike(f'%{search}%'),
                Location.name.ilike(f'%{search}%')
            )
        )

    # Type filter
    if location_type:
        query = query.filter(Location.location_type == location_type)

    # Active/Inactive filter
    if active_filter == 'active':
        query = query.filter(Location.active == True)
    elif active_filter == 'inactive':
        query = query.filter(Location.active == False)

    # Pagination
    pagination = query.order_by(Location.location_type, Location.code).paginate(
        page=page, per_page=50, error_out=False
    )
    locations = pagination.items

    # Get inventory count for each location
    locations_with_inventory = []
    for location in locations:
        inventory_count = db.session.query(func.count(InventoryLevel.id)).filter(
            InventoryLevel.location_id == location.id,
            InventoryLevel.quantity > 0
        ).scalar() or 0

        total_qty = db.session.query(func.sum(InventoryLevel.quantity)).filter(
            InventoryLevel.location_id == location.id
        ).scalar() or 0

        bin_count = Bin.query.filter_by(location_id=location.id, active=True).count()

        locations_with_inventory.append((location, inventory_count, total_qty, bin_count))

    return render_template('locations/index.html',
                          locations=locations_with_inventory,
                          pagination=pagination,
                          search=search,
                          location_type=location_type,
                          active_filter=active_filter)


@bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    """Create new location"""
    if request.method == 'POST':
        try:
            location = Location(
                code=request.form['code'].strip().upper(),
                name=request.form['name'].strip(),
                location_type=request.form['location_type'],
                zone=request.form.get('zone', '').strip(),
                active=request.form.get('active') == 'on'
            )

            db.session.add(location)
            db.session.commit()

            flash(f'Location "{location.code}" created successfully!', 'success')
            return redirect(url_for('locations.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating location: {str(e)}', 'danger')

    return render_template('locations/new.html')


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    """Edit location"""
    location = Location.query.get_or_404(id)

    if request.method == 'POST':
        try:
            location.code = request.form['code'].strip().upper()
            location.name = request.form['name'].strip()
            location.location_type = request.form['location_type']
            location.zone = request.form.get('zone', '').strip()
            location.active = request.form.get('active') == 'on'

            db.session.commit()

            flash(f'Location "{location.code}" updated successfully!', 'success')
            return redirect(url_for('locations.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating location: {str(e)}', 'danger')

    return render_template('locations/edit.html', location=location)


@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    """Delete location"""
    location = Location.query.get_or_404(id)

    # Check if location has inventory
    has_inventory = InventoryLevel.query.filter_by(location_id=id).first() is not None

    if has_inventory:
        flash(f'Cannot delete location "{location.code}" - it has inventory records.', 'danger')
        return redirect(url_for('locations.index'))

    try:
        # Delete associated bins first
        Bin.query.filter_by(location_id=id).delete()
        db.session.delete(location)
        db.session.commit()
        flash(f'Location "{location.code}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting location: {str(e)}', 'danger')

    return redirect(url_for('locations.index'))


@bp.route('/<int:id>/bins')
@login_required
def bins(id):
    """View bins for a location"""
    location = Location.query.get_or_404(id)
    bins = Bin.query.filter_by(location_id=id).order_by(Bin.bin_code).all()

    # Get inventory count for each bin
    bins_with_inventory = []
    for bin in bins:
        inventory_count = db.session.query(func.count(InventoryLevel.id)).filter(
            InventoryLevel.bin_id == bin.id,
            InventoryLevel.quantity > 0
        ).scalar() or 0

        total_qty = db.session.query(func.sum(InventoryLevel.quantity)).filter(
            InventoryLevel.bin_id == bin.id
        ).scalar() or 0

        bins_with_inventory.append((bin, inventory_count, total_qty))

    return render_template('locations/bins.html',
                          location=location,
                          bins=bins_with_inventory)


@bp.route('/<int:id>/bins/new', methods=['GET', 'POST'])
@login_required
def new_bin(id):
    """Create new bin for location"""
    location = Location.query.get_or_404(id)

    if request.method == 'POST':
        try:
            bin = Bin(
                location_id=id,
                bin_code=request.form['bin_code'].strip().upper(),
                description=request.form.get('description', '').strip(),
                active=request.form.get('active') == 'on'
            )

            db.session.add(bin)
            db.session.commit()

            flash(f'Bin "{bin.bin_code}" created successfully!', 'success')
            return redirect(url_for('locations.bins', id=id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating bin: {str(e)}', 'danger')

    return render_template('locations/new_bin.html', location=location)


@bp.route('/bins/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_bin(id):
    """Edit bin"""
    bin = Bin.query.get_or_404(id)
    location = bin.location

    if request.method == 'POST':
        try:
            bin.bin_code = request.form['bin_code'].strip().upper()
            bin.description = request.form.get('description', '').strip()
            bin.active = request.form.get('active') == 'on'

            db.session.commit()

            flash(f'Bin "{bin.bin_code}" updated successfully!', 'success')
            return redirect(url_for('locations.bins', id=location.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating bin: {str(e)}', 'danger')

    return render_template('locations/edit_bin.html', bin=bin, location=location)


@bp.route('/bins/<int:id>/delete', methods=['POST'])
@login_required
def delete_bin(id):
    """Delete bin"""
    bin = Bin.query.get_or_404(id)
    location_id = bin.location_id

    # Check if bin has inventory
    has_inventory = InventoryLevel.query.filter_by(bin_id=id).first() is not None

    if has_inventory:
        flash(f'Cannot delete bin "{bin.bin_code}" - it has inventory records.', 'danger')
        return redirect(url_for('locations.bins', id=location_id))

    try:
        db.session.delete(bin)
        db.session.commit()
        flash(f'Bin "{bin.bin_code}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting bin: {str(e)}', 'danger')

    return redirect(url_for('locations.bins', id=location_id))


@bp.route('/api/bins/<int:location_id>')
@login_required
def api_bins(location_id):
    """API endpoint to get bins for a location"""
    bins = Bin.query.filter_by(location_id=location_id, active=True).order_by(Bin.bin_code).all()
    return jsonify([{'id': b.id, 'bin_code': b.bin_code, 'description': b.description} for b in bins])


@bp.route('/api/available-quantity')
@login_required
def api_available_quantity():
    """API endpoint to get available quantity for item/material at location (sum of all bins)"""
    item_type = request.args.get('item_type')  # 'material' or 'item'
    item_id = request.args.get('item_id', type=int)
    location_id = request.args.get('location_id', type=int)

    if not item_type or not item_id or not location_id:
        return jsonify({'error': 'Missing required parameters'}), 400

    # Query inventory levels at location (all bins)
    query = InventoryLevel.query.filter_by(location_id=location_id)

    if item_type == 'material':
        query = query.filter_by(material_id=item_id)
    elif item_type == 'item':
        query = query.filter_by(item_id=item_id)
    else:
        return jsonify({'error': 'Invalid item type'}), 400

    # Get all inventory records at this location and sum quantities
    inventories = query.all()
    total_quantity = sum(inv.quantity for inv in inventories)

    # Build bin breakdown list
    bins_breakdown = []
    for inv in inventories:
        if inv.quantity > 0:
            bin_info = {
                'bin_code': inv.bin.bin_code if inv.bin else 'No Bin',
                'bin_description': inv.bin.description if inv.bin else '',
                'quantity': inv.quantity
            }
            bins_breakdown.append(bin_info)

    if inventories and total_quantity > 0:
        # Get the item/material name and UOM from first record
        first_inv = inventories[0]
        if item_type == 'material':
            name = first_inv.material.name
            uom = first_inv.material.unit_of_measure
        else:
            name = first_inv.item.name
            uom = first_inv.item.unit_of_measure

        return jsonify({
            'available': True,
            'quantity': total_quantity,
            'name': name,
            'uom': uom,
            'bins': bins_breakdown
        })
    else:
        # Get name/uom even when no stock
        if item_type == 'material':
            from models import Material
            material = Material.query.get(item_id)
            name = material.name if material else 'Unknown'
            uom = material.unit_of_measure if material else ''
        else:
            from models import Item
            item = Item.query.get(item_id)
            name = item.name if item else 'Unknown'
            uom = item.unit_of_measure if item else ''

        return jsonify({
            'available': False,
            'quantity': 0,
            'name': name,
            'uom': uom,
            'bins': []
        })


@bp.route('/api/bins-with-stock')
@login_required
def api_bins_with_stock():
    """API endpoint to get bins with stock quantities for a specific item/material at a location"""
    item_type = request.args.get('item_type')  # 'material' or 'item'
    item_id = request.args.get('item_id', type=int)
    location_id = request.args.get('location_id', type=int)

    if not item_type or not item_id or not location_id:
        return jsonify({'error': 'Missing required parameters'}), 400

    # Query inventory levels for this item/material at this location
    query = InventoryLevel.query.filter_by(location_id=location_id)

    if item_type == 'material':
        query = query.filter_by(material_id=item_id)
    elif item_type == 'item':
        query = query.filter_by(item_id=item_id)
    else:
        return jsonify({'error': 'Invalid item type'}), 400

    # Get all inventory levels with quantity > 0
    inventory_levels = query.filter(InventoryLevel.quantity > 0).all()

    bins_data = []
    for inv in inventory_levels:
        if inv.bin:
            bins_data.append({
                'id': inv.bin.id,
                'bin_code': inv.bin.bin_code,
                'description': inv.bin.description,
                'quantity': inv.quantity
            })

    return jsonify(bins_data)


@bp.route('/bins/import', methods=['GET', 'POST'])
@login_required
def import_bins():
    """Import bins from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded.', 'danger')
            return redirect(url_for('locations.import_bins'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('locations.import_bins'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('locations.import_bins'))

        try:
            wb = load_workbook(file)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            imported = 0
            updated = 0
            errors = []

            for idx, row in enumerate(rows, start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    location_code = str(row[0]).strip().upper()
                    bin_code = str(row[1]).strip().upper() if row[1] else None
                    description = str(row[2] or '').strip() if len(row) > 2 else ''
                    active = str(row[3] or 'Yes').lower() in ['yes', 'true', '1', 'active'] if len(row) > 3 else True

                    if not bin_code:
                        errors.append(f"Row {idx}: Bin code is required")
                        continue

                    # Find the location
                    location = Location.query.filter_by(code=location_code).first()
                    if not location:
                        errors.append(f"Row {idx}: Location '{location_code}' not found")
                        continue

                    # Check if bin already exists for this location
                    existing_bin = Bin.query.filter_by(
                        location_id=location.id,
                        bin_code=bin_code
                    ).first()

                    if existing_bin:
                        # Update existing bin
                        existing_bin.description = description
                        existing_bin.active = active
                        updated += 1
                    else:
                        # Create new bin
                        new_bin = Bin(
                            location_id=location.id,
                            bin_code=bin_code,
                            description=description,
                            active=active
                        )
                        db.session.add(new_bin)
                        imported += 1

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            db.session.commit()

            if errors:
                flash(f'Import completed with errors. Imported: {imported}, Updated: {updated}. Errors: {"; ".join(errors[:5])}{"..." if len(errors) > 5 else ""}', 'warning')
            else:
                flash(f'Import successful! Imported: {imported}, Updated: {updated}.', 'success')

            return redirect(url_for('locations.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error importing file: {str(e)}', 'danger')

    # Get all locations for reference
    locations = Location.query.filter_by(active=True).order_by(Location.code).all()
    return render_template('locations/import_bins.html', locations=locations)


@bp.route('/bins/export')
@login_required
def export_bins():
    """Export all bins to Excel"""
    # Query all bins with their locations
    bins = db.session.query(Bin, Location).join(Location).order_by(Location.code, Bin.bin_code).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bins"

    # Headers with styling
    headers = ['Location Code', 'Bin Code', 'Description', 'Active', 'Location Name', 'Location Type']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, (bin, location) in enumerate(bins, start=2):
        ws.cell(row=row_idx, column=1, value=location.code)
        ws.cell(row=row_idx, column=2, value=bin.bin_code)
        ws.cell(row=row_idx, column=3, value=bin.description or '')
        ws.cell(row=row_idx, column=4, value='Yes' if bin.active else 'No')
        ws.cell(row=row_idx, column=5, value=location.name)
        ws.cell(row=row_idx, column=6, value=location.location_type)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"bins_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/bins/template')
@login_required
def download_bins_template():
    """Download Excel template for bins import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bins Template"

    # Headers with styling
    headers = ['Location Code', 'Bin Code', 'Description', 'Active']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Sample data rows
    sample_data = [
        ['WH-01', 'A-01', 'Aisle A, Row 1', 'Yes'],
        ['WH-01', 'A-02', 'Aisle A, Row 2', 'Yes'],
        ['WH-01', 'B-01', 'Aisle B, Row 1', 'Yes'],
        ['WH-02', 'C-01', 'Zone C, Shelf 1', 'Yes'],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='bins_import_template.xlsx'
    )
