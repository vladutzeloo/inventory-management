"""
Excel Import routes - Bulk stock imports
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Material, Item, Location, Bin, InventoryLevel, Batch, Category, Provider, Client
from datetime import datetime
from werkzeug.utils import secure_filename
from io import BytesIO
from sqlalchemy.exc import IntegrityError
import os

bp = Blueprint('imports', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@bp.route('/')
@login_required
def index():
    """Show import page"""
    return render_template('imports/index.html')


@bp.route('/upload', methods=['POST'])
@login_required
def upload():
    """Process Excel file upload"""
    if 'file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('imports.index'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('imports.index'))

    if not allowed_file(file.filename):
        flash('Invalid file type. Please upload .xlsx or .xls file', 'danger')
        return redirect(url_for('imports.index'))

    try:
        # Try importing openpyxl
        try:
            import openpyxl
        except ImportError:
            flash('Excel import requires openpyxl library. Please install it: pip install openpyxl', 'danger')
            return redirect(url_for('imports.index'))

        # Read Excel file
        workbook = openpyxl.load_workbook(file, read_only=True)
        sheet = workbook.active

        results = {
            'created_materials': 0,
            'created_items': 0,
            'created_categories': 0,
            'created_providers': 0,
            'created_clients': 0,
            'created_bins': 0,
            'updated_stock': 0,
            'errors': []
        }

        # Skip header row
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                # Expected columns: Type, Name, Category, Provider/Client, Location, Bin, Quantity, UOM, Cost, Batch Number, Supplier Batch/IO Number, Diameter, Width, Length, Height
                item_type = str(row[0]).strip().lower() if row[0] else None
                name = str(row[1]).strip() if row[1] else None
                category_name = str(row[2]).strip() if row[2] else None
                org_name = str(row[3]).strip() if row[3] else None  # Provider for materials, Client for items
                location_code = str(row[4]).strip().upper() if row[4] else None
                bin_code = str(row[5]).strip().upper() if row[5] else None

                # Safe float conversion for quantity
                quantity = 0
                if row[6] not in [None, '']:
                    try:
                        quantity = float(row[6])
                    except (ValueError, TypeError):
                        quantity = 0

                uom = str(row[7]).strip() if row[7] else 'PCS'

                # Safe float conversion for cost
                cost = 0
                if row[8] not in [None, '']:
                    try:
                        cost = float(row[8])
                    except (ValueError, TypeError):
                        cost = 0

                # NEW: Batch information
                batch_number = str(row[9]).strip() if row[9] not in [None, ''] else None
                supplier_batch_or_io = str(row[10]).strip() if row[10] not in [None, ''] else None  # Supplier batch for materials, IO number for items

                # Safe float conversion for dimensions
                diameter = None
                if len(row) > 11 and row[11] not in [None, '']:
                    try:
                        diameter = float(row[11])
                    except (ValueError, TypeError):
                        pass

                width = None
                if len(row) > 12 and row[12] not in [None, '']:
                    try:
                        width = float(row[12])
                    except (ValueError, TypeError):
                        pass

                length = None
                if len(row) > 13 and row[13] not in [None, '']:
                    try:
                        length = float(row[13])
                    except (ValueError, TypeError):
                        pass

                height = None
                if len(row) > 14 and row[14] not in [None, '']:
                    try:
                        height = float(row[14])
                    except (ValueError, TypeError):
                        pass

                if not item_type or not name or not location_code:
                    results['errors'].append(f'Row {row_num}: Missing required fields (Type, Name, or Location)')
                    continue

                # Generate batch number if not provided
                if not batch_number:
                    batch_number = f"BATCH-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{row_num}"

                # Get or create category
                category = None
                if category_name:
                    category = Category.query.filter_by(name=category_name, category_type=item_type).first()
                    if not category:
                        try:
                            category = Category(name=category_name, category_type=item_type, active=True)
                            db.session.add(category)
                            db.session.flush()
                            results['created_categories'] += 1
                        except IntegrityError:
                            # Category might have been created by another process or in a previous row
                            # Rollback this flush and query again
                            db.session.rollback()
                            category = Category.query.filter_by(name=category_name, category_type=item_type).first()
                            if not category:
                                # If still not found, this is a real error - skip this row
                                results['errors'].append(f'Row {row_num}: Failed to create/find category "{category_name}"')
                                continue

                # Get or create location
                location = Location.query.filter_by(code=location_code).first()
                if not location:
                    results['errors'].append(f'Row {row_num}: Location {location_code} not found')
                    continue

                # Get or create bin
                bin_obj = None
                if bin_code:
                    bin_obj = Bin.query.filter_by(location_id=location.id, bin_code=bin_code).first()
                    if not bin_obj:
                        bin_obj = Bin(location_id=location.id, bin_code=bin_code, active=True)
                        db.session.add(bin_obj)
                        db.session.flush()
                        results['created_bins'] += 1

                # Process material or item
                if item_type == 'material':
                    # Get or create provider
                    provider = None
                    if org_name:
                        provider = Provider.query.filter_by(name=org_name).first()
                        if not provider:
                            provider = Provider(
                                name=org_name,
                                code=org_name[:10].upper().replace(' ', '_'),
                                active=True
                            )
                            db.session.add(provider)
                            db.session.flush()
                            results['created_providers'] += 1

                    # Get or create material
                    material = Material.query.filter_by(name=name).first()
                    if not material:
                        material = Material(
                            name=name,
                            unit_of_measure=uom,
                            category_id=category.id if category else None,
                            provider_id=provider.id if provider else None,
                            diameter=diameter,
                            width=width,
                            length=length,
                            height=height,
                            active=True
                        )
                        db.session.add(material)
                        db.session.flush()
                        results['created_materials'] += 1
                    else:
                        # Update dimensions if material exists
                        if diameter is not None:
                            material.diameter = diameter
                        if width is not None:
                            material.width = width
                        if length is not None:
                            material.length = length
                        if height is not None:
                            material.height = height

                    # Create batch with batch number
                    batch = Batch(
                        batch_number=batch_number,
                        material_id=material.id,
                        location_id=location.id,
                        bin_id=bin_obj.id if bin_obj else None,
                        quantity_original=quantity,
                        quantity_available=quantity,
                        cost_per_unit=cost,
                        supplier_batch_number=supplier_batch_or_io,
                        received_date=datetime.utcnow(),
                        status='active'
                    )
                    db.session.add(batch)

                    # Update inventory level
                    inv = InventoryLevel.query.filter_by(
                        material_id=material.id,
                        location_id=location.id,
                        bin_id=bin_obj.id if bin_obj else None
                    ).first()
                    if inv:
                        inv.quantity += quantity
                    else:
                        inv = InventoryLevel(
                            material_id=material.id,
                            location_id=location.id,
                            bin_id=bin_obj.id if bin_obj else None,
                            quantity=quantity
                        )
                        db.session.add(inv)
                    results['updated_stock'] += 1

                elif item_type == 'item':
                    # Get or create client
                    client = None
                    if org_name:
                        client = Client.query.filter_by(name=org_name).first()
                        if not client:
                            client = Client(
                                name=org_name,
                                code=org_name[:10].upper().replace(' ', '_'),
                                active=True
                            )
                            db.session.add(client)
                            db.session.flush()
                            results['created_clients'] += 1

                    # Get or create item
                    item = Item.query.filter_by(name=name).first()
                    if not item:
                        item = Item(
                            name=name,
                            unit_of_measure=uom,
                            category_id=category.id if category else None,
                            client_id=client.id if client else None,
                            diameter=diameter,
                            width=width,
                            length=length,
                            height=height,
                            active=True
                        )
                        db.session.add(item)
                        db.session.flush()
                        results['created_items'] += 1
                    else:
                        # Update dimensions if item exists
                        if diameter is not None:
                            item.diameter = diameter
                        if width is not None:
                            item.width = width
                        if length is not None:
                            item.length = length
                        if height is not None:
                            item.height = height

                    # Create batch with batch number
                    batch = Batch(
                        batch_number=batch_number,
                        item_id=item.id,
                        location_id=location.id,
                        bin_id=bin_obj.id if bin_obj else None,
                        quantity_original=quantity,
                        quantity_available=quantity,
                        cost_per_unit=cost,
                        po_number=supplier_batch_or_io,  # Internal order number for items
                        received_date=datetime.utcnow(),
                        status='active'
                    )
                    db.session.add(batch)

                    # Update inventory level
                    inv = InventoryLevel.query.filter_by(
                        item_id=item.id,
                        location_id=location.id,
                        bin_id=bin_obj.id if bin_obj else None
                    ).first()
                    if inv:
                        inv.quantity += quantity
                    else:
                        inv = InventoryLevel(
                            item_id=item.id,
                            location_id=location.id,
                            bin_id=bin_obj.id if bin_obj else None,
                            quantity=quantity
                        )
                        db.session.add(inv)
                    results['updated_stock'] += 1

                else:
                    results['errors'].append(f'Row {row_num}: Invalid type "{item_type}". Must be "material" or "item"')

            except Exception as e:
                results['errors'].append(f'Row {row_num}: {str(e)}')

        # Commit all changes
        db.session.commit()

        # Show results
        flash(f'Import completed! Materials: {results["created_materials"]}, Items: {results["created_items"]}, Stock updated: {results["updated_stock"]}', 'success')
        if results['created_categories']:
            flash(f'Created {results["created_categories"]} new categories', 'info')
        if results['created_providers']:
            flash(f'Created {results["created_providers"]} new providers', 'info')
        if results['created_clients']:
            flash(f'Created {results["created_clients"]} new clients', 'info')
        if results['created_bins']:
            flash(f'Created {results["created_bins"]} new bins', 'info')
        for error in results['errors'][:5]:  # Show first 5 errors
            flash(f'Error: {error}', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {str(e)}', 'danger')

    return redirect(url_for('imports.index'))


@bp.route('/download-template')
@login_required
def download_template():
    """Download Excel template for bulk stock import"""
    try:
        from openpyxl import Workbook
    except ImportError:
        flash('Excel export requires openpyxl library. Please install it: pip install openpyxl', 'danger')
        return redirect(url_for('imports.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Import Template"

    # Headers
    headers = ['Type', 'Name', 'Category', 'Provider/Client', 'Location', 'Bin',
               'Quantity', 'UOM', 'Cost', 'Batch Number', 'Supplier Batch/IO Number',
               'Diameter', 'Width', 'Length', 'Height']
    ws.append(headers)

    # Sample data - Material examples
    ws.append(['material', 'Steel Plate 5mm', 'Metals', 'ABC Steel Co', 'WH-01', 'A-01',
               150, 'KG', 25.50, 'BATCH-MAT-001', 'SUP-12345', '', 1000, 2000, 5])
    ws.append(['material', 'Copper Wire 2.5mm', 'Metals', 'XYZ Suppliers', 'WH-01', 'B-03',
               500, 'M', 2.30, 'BATCH-MAT-002', 'SUP-67890', 2.5, '', '', ''])
    ws.append(['material', 'Aluminum Rod 10mm', 'Metals', 'ABC Steel Co', 'WH-01', 'A-02',
               200, 'PCS', 15.00, 'BATCH-MAT-003', 'SUP-11111', 10, '', 3000, ''])

    # Sample data - Item examples
    ws.append(['item', 'Widget A1000', 'Electronics', 'ClientCo Inc', 'WH-02', 'C-05',
               100, 'PCS', 150.00, 'BATCH-FG-001', 'IO-2025-001', '', 50, 100, 30])
    ws.append(['item', 'Assembly B2000', 'Assemblies', 'TechCorp Ltd', 'WH-02', 'C-06',
               50, 'SET', 299.99, 'BATCH-FG-002', 'IO-2025-002', '', 200, 150, 75])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='stock_import_template.xlsx')
