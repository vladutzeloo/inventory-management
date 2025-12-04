"""
Items master data management routes
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Item, InventoryLevel, Category, Client, Batch, Location, Bin
from sqlalchemy import func, or_
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime

bp = Blueprint('items', __name__)


@bp.route('/')
@login_required
def index():
    """List all items"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    category = request.args.get('category', '', type=str)
    client_id = request.args.get('client_id', '', type=str)
    unit_of_measure = request.args.get('unit_of_measure', '', type=str)
    active_filter = request.args.get('active', '', type=str)

    query = Item.query

    # Search filter
    if search:
        query = query.filter(
            or_(
                Item.name.ilike(f'%{search}%'),
                Item.description.ilike(f'%{search}%')
            )
        )

    # Category filter
    if category:
        query = query.filter(Item.category == category)

    # Client filter
    if client_id:
        query = query.filter(Item.client_id == int(client_id))

    # Unit of Measure filter
    if unit_of_measure:
        query = query.filter(Item.unit_of_measure.ilike(f'%{unit_of_measure}%'))

    # Active/Inactive filter
    if active_filter == 'active':
        query = query.filter(Item.active == True)
    elif active_filter == 'inactive':
        query = query.filter(Item.active == False)

    # Get all categories for filter dropdown
    categories = db.session.query(Item.category).filter(
        Item.category.isnot(None),
        Item.category != ''
    ).distinct().order_by(Item.category).all()
    categories = [c[0] for c in categories]

    # Get all active clients for filter dropdown
    from models import Client
    clients = Client.query.filter_by(active=True).order_by(Client.name).all()

    # Pagination
    pagination = query.order_by(Item.name).paginate(
        page=page, per_page=50, error_out=False
    )
    items = pagination.items

    # Get current stock for each item
    items_with_stock = []
    for item in items:
        stock_qty = db.session.query(func.sum(InventoryLevel.quantity)).filter(
            InventoryLevel.item_id == item.id
        ).scalar() or 0
        items_with_stock.append((item, stock_qty))

    return render_template('items/index.html',
                          items=items_with_stock,
                          pagination=pagination,
                          search=search,
                          category=category,
                          categories=categories,
                          client_id=client_id,
                          clients=clients,
                          unit_of_measure=unit_of_measure,
                          active_filter=active_filter)


@bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    """Create new item"""
    if request.method == 'POST':
        try:
            # Get category_id and client_id from form
            category_id = request.form.get('category_id')
            if category_id:
                category_id = int(category_id) if category_id else None

            client_id = request.form.get('client_id')
            if client_id:
                client_id = int(client_id) if client_id else None

            item = Item(
                name=request.form['name'].strip(),
                description=request.form.get('description', '').strip(),
                category=request.form.get('category', '').strip(),  # Keep legacy field for backwards compatibility
                category_id=category_id,
                client_id=client_id,
                unit_of_measure=request.form['unit_of_measure'].strip(),
                reorder_level=float(request.form.get('reorder_level', 0)),
                reorder_quantity=float(request.form.get('reorder_quantity', 0)),
                diameter=float(request.form.get('diameter')) if request.form.get('diameter') else None,
                length=float(request.form.get('length')) if request.form.get('length') else None,
                width=float(request.form.get('width')) if request.form.get('width') else None,
                height=float(request.form.get('height')) if request.form.get('height') else None,
                active=request.form.get('active') == 'on'
            )

            db.session.add(item)
            db.session.commit()

            flash(f'Item "{item.name}" created successfully!', 'success')
            return redirect(url_for('items.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating item: {str(e)}', 'danger')

    # Get active item categories and clients
    categories = Category.query.filter_by(category_type='item', active=True).order_by(Category.name).all()
    clients = Client.query.filter_by(active=True).order_by(Client.name).all()

    return render_template('items/new.html', categories=categories, clients=clients)


@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    """Edit item"""
    item = Item.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # Get category_id and client_id from form
            category_id = request.form.get('category_id')
            if category_id:
                category_id = int(category_id) if category_id else None

            client_id = request.form.get('client_id')
            if client_id:
                client_id = int(client_id) if client_id else None

            item.name = request.form['name'].strip()
            item.description = request.form.get('description', '').strip()
            item.category = request.form.get('category', '').strip()  # Keep legacy field
            item.category_id = category_id
            item.client_id = client_id
            item.unit_of_measure = request.form['unit_of_measure'].strip()
            item.reorder_level = float(request.form.get('reorder_level', 0))
            item.reorder_quantity = float(request.form.get('reorder_quantity', 0))
            item.diameter = float(request.form.get('diameter')) if request.form.get('diameter') else None
            item.length = float(request.form.get('length')) if request.form.get('length') else None
            item.width = float(request.form.get('width')) if request.form.get('width') else None
            item.height = float(request.form.get('height')) if request.form.get('height') else None
            item.active = request.form.get('active') == 'on'

            db.session.commit()

            flash(f'Item "{item.name}" updated successfully!', 'success')
            return redirect(url_for('items.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error updating item: {str(e)}', 'danger')

    # Get active item categories and clients
    categories = Category.query.filter_by(category_type='item', active=True).order_by(Category.name).all()
    clients = Client.query.filter_by(active=True).order_by(Client.name).all()

    return render_template('items/edit.html', item=item, categories=categories, clients=clients)


@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    """Delete item"""
    item = Item.query.get_or_404(id)

    # Check if item has inventory
    has_inventory = InventoryLevel.query.filter_by(item_id=id).first() is not None

    if has_inventory:
        flash(f'Cannot delete item "{item.name}" - it has inventory records.', 'danger')
        return redirect(url_for('items.index'))

    try:
        db.session.delete(item)
        db.session.commit()
        flash(f'Item "{item.name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting item: {str(e)}', 'danger')

    return redirect(url_for('items.index'))


@bp.route('/export')
@login_required
def export():
    """Export items to Excel"""
    items = Item.query.order_by(Item.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    # Headers
    headers = ['Name', 'Description', 'Category', 'Unit of Measure',
               'Reorder Level', 'Reorder Quantity', 'Diameter', 'Length', 'Width', 'Height', 'Active']
    ws.append(headers)

    # Data
    for item in items:
        ws.append([
            item.name,
            item.description,
            item.category,
            item.unit_of_measure,
            item.reorder_level,
            item.reorder_quantity,
            item.diameter if item.diameter else '',
            item.length if item.length else '',
            item.width if item.width else '',
            item.height if item.height else '',
            'Yes' if item.active else 'No'
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_data():
    """Import items from Excel"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded.', 'danger')
            return redirect(url_for('items.import_data'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('items.import_data'))

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('items.import_data'))

        try:
            wb = load_workbook(file)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            imported = 0
            updated = 0
            errors = []

            for idx, row in enumerate(rows, start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    name = str(row[0]).strip()
                    description = str(row[1] or '').strip()
                    category = str(row[2] or '').strip()
                    unit_of_measure = str(row[3] or 'PCS').strip()
                    reorder_level = float(row[4] or 0)
                    reorder_quantity = float(row[5] or 0)
                    active = str(row[6] or 'Yes').lower() in ['yes', 'true', '1', 'active']

                    # Batch information (optional)
                    batch_number = str(row[7]).strip() if len(row) > 7 and row[7] else None
                    internal_order_number = str(row[8]).strip() if len(row) > 8 and row[8] else None
                    location_code = str(row[9]).strip() if len(row) > 9 and row[9] else None
                    bin_code = str(row[10]).strip() if len(row) > 10 and row[10] else None
                    quantity = float(row[11]) if len(row) > 11 and row[11] else None
                    cost_per_unit = float(row[12]) if len(row) > 12 and row[12] else None
                    received_date_str = str(row[13]).strip() if len(row) > 13 and row[13] else None

                    # Check if item exists
                    item = Item.query.filter_by(name=name).first()

                    if item:
                        # Update existing
                        item.description = description
                        item.category = category
                        item.unit_of_measure = unit_of_measure
                        item.reorder_level = reorder_level
                        item.reorder_quantity = reorder_quantity
                        item.active = active
                        updated += 1
                    else:
                        # Create new
                        item = Item(
                            name=name,
                            description=description,
                            category=category,
                            unit_of_measure=unit_of_measure,
                            reorder_level=reorder_level,
                            reorder_quantity=reorder_quantity,
                            active=active
                        )
                        db.session.add(item)
                        imported += 1

                    # Flush to get item ID for batch creation
                    db.session.flush()

                    # Create batch if batch information is provided
                    if batch_number and location_code and quantity and cost_per_unit:
                        # Find location
                        location = Location.query.filter_by(code=location_code).first()
                        if not location:
                            errors.append(f"Row {idx}: Location '{location_code}' not found")
                            continue

                        # Find bin if provided
                        bin_obj = None
                        if bin_code:
                            bin_obj = Bin.query.filter_by(location_id=location.id, bin_code=bin_code).first()
                            if not bin_obj:
                                errors.append(f"Row {idx}: Bin '{bin_code}' not found in location '{location_code}'")
                                continue

                        # Parse received date
                        received_date = datetime.utcnow()
                        if received_date_str:
                            try:
                                # Try different date formats
                                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                                    try:
                                        received_date = datetime.strptime(received_date_str, fmt)
                                        break
                                    except ValueError:
                                        continue
                            except:
                                pass  # Use default date if parsing fails

                        # Check if batch already exists
                        existing_batch = Batch.query.filter_by(batch_number=batch_number).first()
                        if not existing_batch:
                            # Create new batch
                            batch = Batch(
                                batch_number=batch_number,
                                item_id=item.id,
                                location_id=location.id,
                                bin_id=bin_obj.id if bin_obj else None,
                                quantity_original=quantity,
                                quantity_available=quantity,
                                cost_per_unit=cost_per_unit,
                                po_number=internal_order_number,  # Using po_number for internal order number
                                received_date=received_date,
                                status='active'
                            )
                            db.session.add(batch)

                            # Update inventory level
                            inventory = InventoryLevel.query.filter_by(
                                item_id=item.id,
                                location_id=location.id,
                                bin_id=bin_obj.id if bin_obj else None
                            ).first()

                            if inventory:
                                inventory.quantity += quantity
                            else:
                                inventory = InventoryLevel(
                                    item_id=item.id,
                                    location_id=location.id,
                                    bin_id=bin_obj.id if bin_obj else None,
                                    quantity=quantity
                                )
                                db.session.add(inventory)

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            db.session.commit()

            if errors:
                flash(f'Import completed with errors. Imported: {imported}, Updated: {updated}. Errors: {"; ".join(errors[:5])}', 'warning')
            else:
                flash(f'Import successful! Imported: {imported}, Updated: {updated}.', 'success')

            return redirect(url_for('items.index'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error importing file: {str(e)}', 'danger')

    return render_template('items/import.html')


@bp.route('/template')
@login_required
def download_template():
    """Download Excel template for items import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Template"

    # Headers
    headers = ['Name', 'Description', 'Category', 'Unit of Measure',
               'Reorder Level', 'Reorder Quantity', 'Active', 'Batch Number',
               'Internal Order Number', 'Location Code', 'Bin Code', 'Quantity',
               'Cost per Unit', 'Received Date']
    ws.append(headers)

    # Sample data
    ws.append(['Widget A', 'Standard widget model A', 'Finished Goods', 'PCS', 50, 100, 'Yes',
               'BATCH-FG-001', 'IO-2025-001', 'WH-001', 'FG-01', 100, 45.00, '2025-01-15'])
    ws.append(['Assembly B', 'Complete assembly model B', 'Assemblies', 'PCS', 25, 50, 'Yes',
               'BATCH-FG-002', 'IO-2025-002', 'WH-001', 'FG-02', 75, 125.50, '2025-01-16'])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='items_template.xlsx')
